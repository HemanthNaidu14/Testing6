import openpyxl
from selenium import webdriver
driver=webdriver.Chrome(executable_path="C:\\Users\\Hemanth Naidu\\chromedriver_win32\\chromedriver.exe")'''
driver.implicitly_wait(10)
driver.maximize_window()
driver.get("https://www.globalsqa.com/samplepagetest")
driver.find_element_by_xpath("//body/div[@id='wrapper']/div[1]/div[2]/div[1]/div[1]/div[2]/div[2]/form[1]/div[6]/label[2]/input[1]").click()'''

driver.implicitly_wait(10)
driver.maximize_window()
driver.get("https://www.globalsqa.com/samplepagetest")
driver.find_element_by_xpath("//body/div[@id='wrapper']/div[1]/div[2]/div[1]/div[1]/div[2]/div[2]/form[1]/div[6]/label[3]/input[1]").click()'''

#script to load workbook from local memory to program

from selenium import webdriver
driver=webdriver.Chrome(executable_path="C:\\Users\\Hemanth Naidu\\chromedriver_win32\\chromedriver.exe")
driver.implicitly_wait(10)
driver.maximize_window()
book=openpyxl.load_workbook("E:\\Hemi\\hem.xlsx")
sheet=book.active
cell=sheet.cell(row=2,column=2)
print(cell.value)
#assign a value in desired cell
sheet.cell(row=1,column=11).value="Hemanth"
print(sheet.cell(row=1,column=11).value
print(sheet.max_row)
print(sheet.max_column)
